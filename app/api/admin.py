import json
import secrets
import string
from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.dependencies import get_current_admin
from app.core.security import get_password_hash
from app.models.models import (
    ClassroomLayout,
    QueueStatus,
    RoundQueueItem,
    Seat,
    SeatRound,
    SeatSelection,
    SeatStatus,
    Semester,
    SpecialSeatRequest,
    SpecialRequestStatus,
    SwapRequest,
    SwapStatus,
    TeamInvite,
    User,
    UserRole,
)
from app.schemas.interaction import (
    ForceSwapRequest,
    ImportedStudentCredential,
    QueueActionResponse,
    ScoreImportResult,
    SeatLockRequest,
    SpecialRequestOut,
    SpecialRequestReview,
    SwapRequestOut,
    SwapReviewRequest,
)
from app.schemas.layout import LayoutCreate, LayoutOut, LayoutUpdate, SeatAdminUpdate, SeatOut
from app.schemas.round import QueueItemOut, RoundCreate, RoundOut, RoundUpdate, SemesterCreate, SemesterOut, SemesterUpdate
from app.schemas.settings import SiteSettingOut, SiteSettingUpdate
from app.schemas.user import StudentScoreOut, UserCreate, UserOut, UserUpdate
from app.services.selection import (
    activate_next_batch,
    apply_lock_to_student,
    finish_round,
    generate_layout_seats,
    get_active_layout,
    get_layout_or_404,
    get_or_create_settings,
    get_round_or_404,
    get_seat_or_404,
    get_student_by_student_id,
    prepare_round,
    skip_current_batch,
    start_makeup_phase,
    swap_students_in_round,
)
from app.services.webhook import send_webhook_if_configured

router = APIRouter(prefix="/admin", tags=["admin"], dependencies=[Depends(get_current_admin)])


def _generate_temporary_password(length: int = 12) -> str:
    alphabet = string.ascii_letters + string.digits
    return "".join(secrets.choice(alphabet) for _ in range(length))


def _build_unique_import_username(base_username: str, existing_usernames: set[str]) -> str:
    if base_username not in existing_usernames:
        return base_username

    suffix = 1
    while f"{base_username}_{suffix}" in existing_usernames:
        suffix += 1
    return f"{base_username}_{suffix}"


def _ensure_unique_active_layout(db: Session, layout_id: int) -> None:
    db.query(ClassroomLayout).filter(ClassroomLayout.id != layout_id).update({"is_active": False}, synchronize_session=False)


def _ensure_unique_active_semester(db: Session, semester_id: int) -> None:
    db.query(Semester).filter(Semester.id != semester_id).update({"is_active": False}, synchronize_session=False)


def _layout_payload(layout: ClassroomLayout, seats: list[Seat]) -> dict:
    return {
        "id": layout.id,
        "name": layout.name,
        "rows": layout.rows,
        "cols": layout.cols,
        "is_active": layout.is_active,
        "seats": [SeatOut.model_validate(seat).model_dump() for seat in seats],
    }


@router.post("/users", response_model=UserOut, summary="Create user")
def create_user(payload: UserCreate, db: Session = Depends(get_db)) -> User:
    if db.query(User).filter(User.username == payload.username).first():
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Username already exists")
    if payload.student_id and db.query(User).filter(User.student_id == payload.student_id).first():
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Student ID already exists")

    user = User(
        username=payload.username,
        full_name=payload.full_name,
        role=payload.role,
        student_id=payload.student_id,
        hashed_password=get_password_hash(payload.password),
        moral_score=payload.moral_score,
        is_active=payload.is_active,
        must_change_password=payload.must_change_password,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return user


@router.get("/users", response_model=list[UserOut], summary="List users")
def list_users(db: Session = Depends(get_db)) -> list[User]:
    return db.query(User).order_by(User.role, User.id).all()


@router.put("/users/{user_id}", response_model=UserOut, summary="Update user")
def update_user(user_id: int, payload: UserUpdate, db: Session = Depends(get_db)) -> User:
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    data = payload.model_dump(exclude_unset=True)
    password = data.pop("password", None)
    if "student_id" in data and data["student_id"]:
        conflict = db.query(User).filter(User.student_id == data["student_id"], User.id != user_id).first()
        if conflict:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Student ID already exists")
    for key, value in data.items():
        setattr(user, key, value)
    if password:
        user.hashed_password = get_password_hash(password)
    db.commit()
    db.refresh(user)
    return user


@router.delete("/users/{user_id}", summary="Delete user")
def delete_user(user_id: int, db: Session = Depends(get_db)) -> dict:
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    db.delete(user)
    db.commit()
    return {"message": "User deleted"}


@router.get("/students", response_model=list[StudentScoreOut], summary="List students and moral scores")
def list_students(db: Session = Depends(get_db)) -> list[User]:
    return db.query(User).filter(User.role == UserRole.student).order_by(User.student_id, User.username).all()


@router.post("/students", response_model=UserOut, summary="Create student")
def create_student(payload: UserCreate, db: Session = Depends(get_db)) -> User:
    if payload.role != UserRole.student:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="This endpoint only creates student accounts")
    return create_user(payload, db)


@router.post("/score-import-excel", response_model=ScoreImportResult, summary="Import moral scores from Excel file")
async def import_scores_excel(file: UploadFile = File(...), db: Session = Depends(get_db)) -> ScoreImportResult:
    content = await file.read()
    try:
        df = pd.read_excel(BytesIO(content))
        # Support basic column names mapping
        # Expecting 'student_id' and 'moral_score' OR '学号' and '德育学分'
        if '学号' in df.columns:
            df = df.rename(columns={'学号': 'student_id', '德育学分': 'moral_score', '姓名': 'full_name'})
        
        # Ensure student_id is string
        df['student_id'] = df['student_id'].astype(str)
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid Excel file") from exc

    updated_count = 0
    ignored_student_ids: list[str] = []
    created_students: list[ImportedStudentCredential] = []
    
    # Pre-fetch all students to optimize queries
    student_ids_in_excel = df['student_id'].tolist()
    students_in_db = db.query(User).filter(User.role == UserRole.student, User.student_id.in_(student_ids_in_excel)).all()
    student_map = {s.student_id: s for s in students_in_db}
    existing_usernames = {username for (username,) in db.query(User.username).all()}
    
    for _, row in df.iterrows():
        student_id = str(row.get('student_id'))
        score = row.get('moral_score')
        full_name = row.get('full_name') if pd.notna(row.get('full_name')) else None
        
        if pd.isna(score):
            ignored_student_ids.append(student_id)
            continue
            
        student = student_map.get(student_id)
        if student:
            student.moral_score = float(score)
            updated_count += 1
        else:
            # Optionally create the student if they don't exist
            if full_name:
                username = _build_unique_import_username(f"stu_{student_id}", existing_usernames)
                temporary_password = _generate_temporary_password()
                new_student = User(
                    username=username,
                    full_name=str(full_name),
                    role=UserRole.student,
                    student_id=student_id,
                    hashed_password=get_password_hash(temporary_password),
                    moral_score=float(score),
                    is_active=True,
                    must_change_password=True,
                )
                db.add(new_student)
                student_map[student_id] = new_student
                existing_usernames.add(username)
                created_students.append(
                    ImportedStudentCredential(
                        student_id=student_id,
                        username=username,
                        temporary_password=temporary_password,
                    )
                )
                updated_count += 1
            else:
                ignored_student_ids.append(student_id)
            
    db.commit()
    return ScoreImportResult(
        updated_count=updated_count,
        ignored_count=len(ignored_student_ids),
        ignored_student_ids=ignored_student_ids,
        created_students=created_students,
    )

@router.post("/score-import", response_model=ScoreImportResult, summary="Import moral scores from JSON file")
async def import_scores(file: UploadFile = File(...), db: Session = Depends(get_db)) -> ScoreImportResult:
    content = await file.read()
    try:
        payload = json.loads(content.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid JSON file") from exc

    students = payload.get("students", [])
    updated_count = 0
    ignored_student_ids: list[str] = []
    for item in students:
        student_id = item.get("student_id")
        score = item.get("moral_score")
        student = (
            db.query(User)
            .filter(User.role == UserRole.student, User.student_id == str(student_id))
            .first()
        )
        if not student or score is None:
            ignored_student_ids.append(str(student_id))
            continue
        student.moral_score = float(score)
        updated_count += 1
    db.commit()
    return ScoreImportResult(
        updated_count=updated_count,
        ignored_count=len(ignored_student_ids),
        ignored_student_ids=ignored_student_ids,
        created_students=[],
    )


@router.post("/students/import-scores", response_model=ScoreImportResult, summary="Import moral scores from JSON file")
async def import_scores_compat(file: UploadFile = File(...), db: Session = Depends(get_db)) -> ScoreImportResult:
    return await import_scores(file, db)


@router.get("/settings", response_model=SiteSettingOut, summary="Get site settings")
def get_settings(db: Session = Depends(get_db)):
    return get_or_create_settings(db)


@router.put("/settings", response_model=SiteSettingOut, summary="Update site settings")
def update_settings(payload: SiteSettingUpdate, db: Session = Depends(get_db)):
    settings_obj = get_or_create_settings(db)
    for key, value in payload.model_dump(exclude_unset=True).items():
        setattr(settings_obj, key, str(value) if key == "webhook_url" and value is not None else value)
    db.commit()
    db.refresh(settings_obj)
    return settings_obj


@router.post("/semesters", response_model=SemesterOut, summary="Create semester")
def create_semester(payload: SemesterCreate, db: Session = Depends(get_db)):
    semester = Semester(name=payload.name, is_active=payload.is_active)
    db.add(semester)
    db.commit()
    db.refresh(semester)
    if semester.is_active:
        _ensure_unique_active_semester(db, semester.id)
        db.commit()
        db.refresh(semester)
    return semester


@router.get("/semesters", response_model=list[SemesterOut], summary="List semesters")
def list_semesters(db: Session = Depends(get_db)):
    return db.query(Semester).order_by(Semester.id.desc()).all()


@router.put("/semesters/{semester_id}", response_model=SemesterOut, summary="Update semester")
def update_semester(semester_id: int, payload: SemesterUpdate, db: Session = Depends(get_db)):
    semester = db.query(Semester).filter(Semester.id == semester_id).first()
    if not semester:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Semester not found")
    for key, value in payload.model_dump(exclude_unset=True).items():
        setattr(semester, key, value)
    db.commit()
    if semester.is_active:
        _ensure_unique_active_semester(db, semester.id)
        db.commit()
    db.refresh(semester)
    return semester


@router.delete("/semesters/{semester_id}", summary="Delete semester")
def delete_semester(semester_id: int, db: Session = Depends(get_db)):
    semester = db.query(Semester).filter(Semester.id == semester_id).first()
    if not semester:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Semester not found")
    db.delete(semester)
    db.commit()
    return {"message": "Semester deleted"}


@router.post("/layouts", response_model=LayoutOut, summary="Create classroom layout")
def create_layout(payload: LayoutCreate, db: Session = Depends(get_db)):
    layout = ClassroomLayout(name=payload.name, rows=payload.rows, cols=payload.cols, is_active=payload.is_active)
    db.add(layout)
    db.commit()
    db.refresh(layout)
    if layout.is_active:
        _ensure_unique_active_layout(db, layout.id)
        db.commit()
    seats = generate_layout_seats(db, layout)
    return _layout_payload(layout, seats)


@router.get("/layouts", response_model=list[LayoutOut], summary="List classroom layouts")
def list_layouts(db: Session = Depends(get_db)):
    layouts = db.query(ClassroomLayout).order_by(ClassroomLayout.id.desc()).all()
    result = []
    for layout in layouts:
        seats = db.query(Seat).filter(Seat.layout_id == layout.id).order_by(Seat.row_index, Seat.col_index).all()
        result.append(_layout_payload(layout, seats))
    return result


@router.put("/layouts/{layout_id}", response_model=LayoutOut, summary="Update classroom layout")
def update_layout(layout_id: int, payload: LayoutUpdate, db: Session = Depends(get_db)):
    layout = get_layout_or_404(db, layout_id)
    data = payload.model_dump(exclude_unset=True)
    rows_changed = "rows" in data and data["rows"] != layout.rows
    cols_changed = "cols" in data and data["cols"] != layout.cols
    for key, value in data.items():
        setattr(layout, key, value)
    if layout.is_active:
        _ensure_unique_active_layout(db, layout.id)
    if rows_changed or cols_changed:
        occupied = (
            db.query(Seat)
            .filter(Seat.layout_id == layout.id)
            .filter((Seat.selected_student_id.is_not(None)) | (Seat.locked_student_id.is_not(None)))
            .count()
        )
        if occupied:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Cannot regenerate occupied layout")
        db.query(Seat).filter(Seat.layout_id == layout.id).delete(synchronize_session=False)
        db.commit()
        seats = generate_layout_seats(db, layout)
    else:
        db.commit()
        seats = db.query(Seat).filter(Seat.layout_id == layout.id).order_by(Seat.row_index, Seat.col_index).all()
    db.refresh(layout)
    return _layout_payload(layout, seats)


@router.delete("/layouts/{layout_id}", summary="Delete classroom layout")
def delete_layout(layout_id: int, db: Session = Depends(get_db)):
    layout = get_layout_or_404(db, layout_id)
    db.query(Seat).filter(Seat.layout_id == layout.id).delete(synchronize_session=False)
    db.delete(layout)
    db.commit()
    return {"message": "Layout deleted"}


@router.put("/seats/{seat_id}", response_model=SeatOut, summary="Update seat tags")
def update_seat(seat_id: int, payload: SeatAdminUpdate, db: Session = Depends(get_db)):
    seat = get_seat_or_404(db, seat_id)
    if payload.tags is not None:
        seat.tags = payload.tags
    db.commit()
    db.refresh(seat)
    return seat


@router.post("/rounds", response_model=RoundOut, summary="Create seat selection round")
def create_round(payload: RoundCreate, db: Session = Depends(get_db)):
    semester = db.query(Semester).filter(Semester.id == payload.semester_id).first()
    if not semester:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Semester not found")
    round_obj = SeatRound(
        semester_id=payload.semester_id,
        name=payload.name,
        status=payload.status,
        ranking_tiebreak=payload.ranking_tiebreak,
        allowed_selection_count=payload.allowed_selection_count,
    )
    db.add(round_obj)
    db.commit()
    db.refresh(round_obj)
    return round_obj


@router.get("/rounds", response_model=list[RoundOut], summary="List seat selection rounds")
def list_rounds(db: Session = Depends(get_db)):
    return db.query(SeatRound).order_by(SeatRound.id.desc()).all()


@router.get("/semesters/{semester_id}/rounds", response_model=list[RoundOut], summary="List rounds under a semester")
def list_semester_rounds(semester_id: int, db: Session = Depends(get_db)):
    return (
        db.query(SeatRound)
        .filter(SeatRound.semester_id == semester_id)
        .order_by(SeatRound.id.desc())
        .all()
    )


@router.put("/rounds/{round_id}", response_model=RoundOut, summary="Update seat selection round")
def update_round(round_id: int, payload: RoundUpdate, db: Session = Depends(get_db)):
    round_obj = get_round_or_404(db, round_id)
    for key, value in payload.model_dump(exclude_unset=True).items():
        setattr(round_obj, key, value)
    db.commit()
    db.refresh(round_obj)
    return round_obj


@router.delete("/rounds/{round_id}", summary="Delete seat selection round")
def delete_round(round_id: int, db: Session = Depends(get_db)):
    round_obj = get_round_or_404(db, round_id)
    db.query(SeatSelection).filter(SeatSelection.round_id == round_id).delete(synchronize_session=False)
    db.query(RoundQueueItem).filter(RoundQueueItem.round_id == round_id).delete(synchronize_session=False)
    db.query(TeamInvite).filter(TeamInvite.round_id == round_id).delete(synchronize_session=False)
    db.query(SwapRequest).filter(SwapRequest.round_id == round_id).delete(synchronize_session=False)
    db.query(SpecialSeatRequest).filter(SpecialSeatRequest.round_id == round_id).delete(synchronize_session=False)
    db.delete(round_obj)
    db.commit()
    return {"message": "Round deleted"}


@router.post("/rounds/{round_id}/prepare", response_model=QueueActionResponse, summary="Prepare round and rebuild ranking queue")
def prepare_round_endpoint(round_id: int, db: Session = Depends(get_db)):
    round_obj = get_round_or_404(db, round_id)
    queue_items = prepare_round(db, round_obj)
    return QueueActionResponse(message="Round prepared", current_student_ids=[str(len(queue_items))])


@router.post("/rounds/{round_id}/open", response_model=QueueActionResponse, summary="Open seat selection site")
def open_round(round_id: int, db: Session = Depends(get_db)):
    round_obj = get_round_or_404(db, round_id)
    round_obj.site_open = True
    db.commit()
    return QueueActionResponse(message="Seat selection site opened")


@router.post("/rounds/{round_id}/open-site", response_model=QueueActionResponse, summary="Open seat selection site")
def open_round_compat(round_id: int, db: Session = Depends(get_db)):
    return open_round(round_id, db)


@router.post("/rounds/{round_id}/close", response_model=QueueActionResponse, summary="Close seat selection site")
def close_round(round_id: int, db: Session = Depends(get_db)):
    round_obj = get_round_or_404(db, round_id)
    round_obj.site_open = False
    db.commit()
    return QueueActionResponse(message="Seat selection site closed")


@router.post("/rounds/{round_id}/close-site", response_model=QueueActionResponse, summary="Close seat selection site")
def close_round_compat(round_id: int, db: Session = Depends(get_db)):
    return close_round(round_id, db)


@router.get("/rounds/{round_id}/queue", response_model=list[QueueItemOut], summary="Get ranked queue for a round")
def get_round_queue(round_id: int, db: Session = Depends(get_db)):
    items = (
        db.query(RoundQueueItem)
        .filter(RoundQueueItem.round_id == round_id)
        .order_by(RoundQueueItem.rank_order)
        .all()
    )
    return [
        {
            "id": item.id,
            "round_id": item.round_id,
            "student_id": item.student_id,
            "student_code": item.student.student_id or item.student.username,
            "student_name": item.student.full_name,
            "rank_order": item.rank_order,
            "phase": item.phase,
            "status": item.status,
            "swap_count": item.swap_count,
        }
        for item in items
    ]


@router.post("/rounds/{round_id}/advance", response_model=QueueActionResponse, summary="Move to next selectable batch")
async def advance_round(round_id: int, db: Session = Depends(get_db)):
    round_obj = get_round_or_404(db, round_id)
    next_items = activate_next_batch(db, round_obj)
    student_codes = [item.student.student_id or item.student.username for item in next_items]
    for item in next_items:
        await send_webhook_if_configured(
            db,
            {
                "event": "turn_to_select",
                "round_id": round_id,
                "student_id": item.student.student_id or item.student.username,
            },
        )
    return QueueActionResponse(message="Advanced to next batch", current_student_ids=student_codes)


@router.post("/rounds/{round_id}/next", response_model=QueueActionResponse, summary="Move to next selectable batch")
async def advance_round_compat(round_id: int, db: Session = Depends(get_db)):
    return await advance_round(round_id, db)


@router.post("/rounds/{round_id}/skip-current", response_model=QueueActionResponse, summary="Skip current selectable batch")
async def skip_current(round_id: int, db: Session = Depends(get_db)):
    round_obj = get_round_or_404(db, round_id)
    skipped_items = skip_current_batch(db, round_obj)
    student_codes = [item.student.student_id or item.student.username for item in skipped_items]
    for item in skipped_items:
        await send_webhook_if_configured(
            db,
            {
                "event": "student_skipped",
                "round_id": round_id,
                "student_id": item.student.student_id or item.student.username,
            },
        )
    return QueueActionResponse(message="Current batch skipped", current_student_ids=student_codes)


@router.post("/rounds/{round_id}/skip", response_model=QueueActionResponse, summary="Skip current selectable batch")
async def skip_current_compat(round_id: int, db: Session = Depends(get_db)):
    return await skip_current(round_id, db)


@router.post("/rounds/{round_id}/start-makeup", response_model=QueueActionResponse, summary="Start makeup phase")
def start_makeup(round_id: int, db: Session = Depends(get_db)):
    round_obj = get_round_or_404(db, round_id)
    count = start_makeup_phase(db, round_obj)
    return QueueActionResponse(message=f"Makeup phase started with {count} pending student(s)")


@router.post("/rounds/{round_id}/finish", response_model=QueueActionResponse, summary="Finish current round")
async def finish_round_endpoint(round_id: int, db: Session = Depends(get_db)):
    round_obj = get_round_or_404(db, round_id)
    finish_round(db, round_obj)
    await send_webhook_if_configured(db, {"event": "round_finished", "round_id": round_id})
    return QueueActionResponse(message="Round finished")


@router.post("/rounds/{round_id}/lock-seat", response_model=QueueActionResponse, summary="Lock or unlock seat for student")
def lock_seat(round_id: int, payload: SeatLockRequest, db: Session = Depends(get_db)):
    round_obj = get_round_or_404(db, round_id)
    seat = get_seat_or_404(db, payload.seat_id)
    if payload.student_id is None:
        seat.locked_student_id = None
        if seat.status == SeatStatus.locked:
            seat.status = SeatStatus.available
            seat.selected_student_id = None
        db.commit()
        return QueueActionResponse(message="Seat unlocked")

    student = get_student_by_student_id(db, payload.student_id)
    apply_lock_to_student(db, round_obj, student, seat)
    return QueueActionResponse(message="Seat locked", current_student_ids=[payload.student_id])


@router.get("/special-requests", response_model=list[SpecialRequestOut], summary="List special seat requests")
def list_special_requests(db: Session = Depends(get_db)):
    return db.query(SpecialSeatRequest).order_by(SpecialSeatRequest.id.desc()).all()


@router.post("/special-requests/{request_id}/review", response_model=SpecialRequestOut, summary="Review special seat request")
def review_special_request(request_id: int, payload: SpecialRequestReview, db: Session = Depends(get_db), admin: User = Depends(get_current_admin)):
    request_obj = db.query(SpecialSeatRequest).filter(SpecialSeatRequest.id == request_id).first()
    if not request_obj:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Special request not found")
    request_obj.review_comment = payload.review_comment
    request_obj.reviewer_id = admin.id
    request_obj.status = SpecialRequestStatus.approved if payload.approve else SpecialRequestStatus.rejected
    if payload.approve and payload.seat_id is not None:
        seat = get_seat_or_404(db, payload.seat_id)
        round_obj = get_round_or_404(db, request_obj.round_id)
        apply_lock_to_student(db, round_obj, request_obj.student, seat)
    else:
        db.commit()
    db.refresh(request_obj)
    return request_obj


@router.get("/swap-requests", response_model=list[SwapRequestOut], summary="List swap requests")
def list_swap_requests(db: Session = Depends(get_db)):
    return db.query(SwapRequest).order_by(SwapRequest.id.desc()).all()


@router.post("/swap-requests/{request_id}/review", response_model=SwapRequestOut, summary="Approve or reject swap request")
def review_swap_request(request_id: int, payload: SwapReviewRequest, db: Session = Depends(get_db)):
    request_obj = db.query(SwapRequest).filter(SwapRequest.id == request_id).first()
    if not request_obj:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Swap request not found")
    request_obj.review_comment = payload.review_comment
    if payload.approve:
        swap_students_in_round(db, request_obj.round_id, request_obj.requester_id, request_obj.target_id, increment_counter=True)
        request_obj.status = SwapStatus.approved
    else:
        request_obj.status = SwapStatus.rejected
        db.commit()
    db.refresh(request_obj)
    return request_obj


@router.post("/rounds/{round_id}/force-swap", response_model=QueueActionResponse, summary="Force swap two students")
def force_swap(round_id: int, payload: ForceSwapRequest, db: Session = Depends(get_db)):
    first_student = get_student_by_student_id(db, payload.first_student_id)
    second_student = get_student_by_student_id(db, payload.second_student_id)
    swap_students_in_round(db, round_id, first_student.id, second_student.id, increment_counter=False)
    return QueueActionResponse(
        message="Seats swapped",
        current_student_ids=[payload.first_student_id, payload.second_student_id],
    )


@router.get("/rounds/{round_id}/export", summary="Export final seat table to Excel")
def export_round(round_id: int, db: Session = Depends(get_db)):
    round_obj = get_round_or_404(db, round_id)
    layout = get_active_layout(db)
    seats = db.query(Seat).filter(Seat.layout_id == layout.id).all()
    selection_map = {
        selection.seat_id: selection
        for selection in db.query(SeatSelection).filter(SeatSelection.round_id == round_obj.id).all()
    }

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "SeatTable"
    sheet.cell(row=1, column=1, value="Seat Code")
    sheet.cell(row=1, column=2, value="Student ID")
    sheet.cell(row=1, column=3, value="Student Name")
    row_no = 2
    for seat in sorted(seats, key=lambda item: (item.row_index, item.col_index)):
        selection = selection_map.get(seat.id)
        student = selection.student if selection else None
        sheet.cell(row=row_no, column=1, value=seat.seat_code)
        sheet.cell(row=row_no, column=2, value=student.student_id if student else "")
        sheet.cell(row=row_no, column=3, value=student.full_name if student else "")
        row_no += 1

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    filename = f"round_{round_id}_seat_table.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
